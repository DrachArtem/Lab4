import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import datetime

# Зчитуємо дані з CSV файлу
try:
    df = pd.read_csv('employees.csv')
except FileNotFoundError:
    print("Помилка: Файл employees.csv не знайдено.")
    exit()

# Створюємо новий Excel файл
wb = Workbook()

# Створюємо аркуш "all" та записуємо всі дані
ws_all = wb.create_sheet('all')
for r_idx, row_data in enumerate(df.values, start=1):
    for c_idx, cell_value in enumerate(row_data, start=1):
        ws_all.cell(row=r_idx, column=c_idx, value=cell_value)

# Створюємо інші аркуші з віковими категоріями
age_categories = {
    "younger_18": (0, 17),
    "18-45": (18, 45),
    "45-70": (46, 70),
    "older_70": (71, 150),  # Максимальний вік
}

for category, (min_age, max_age) in age_categories.items():
    ws = wb.create_sheet(category)
    ws.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])
    idx = 1
    for _, row in df.iterrows():
        birthdate = datetime.datetime.strptime(row['Дата народження'], '%Y-%m-%d')
        age = datetime.datetime.now().year - birthdate.year
        if min_age <= age <= max_age:
            ws.append([idx, row['Прізвище'], row['Ім’я'], row['По батькові'], row['Дата народження'], age])
            idx += 1

# Видаляємо створений за замовчуванням аркуш
del wb['Sheet']

# Зберігаємо результати в Excel файл
try:
    wb.save('employees.xlsx')
    print("Ok")
except PermissionError:
    print("Помилка: Немає прав для створення файлу XLSX.")
except Exception as e:
    print(f"Помилка: {e}")
